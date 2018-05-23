import openpyxl
import psycopg2
import shutil
import re
import time
import datetime
import os
from ConfigParser import ConfigParser

now = datetime.datetime.now()

t0 = time.clock()

#path_files = "/var/www/python/read_excel"
path_files = "/home/cmasiva"
config = ConfigParser()
config.read(path_files+"/.env")

user = 'postgres'
db = "contactosms"
host = "10.221.16.88"
pw = "123"

#user = config.get("DATABASE", "user")
#db = config.get("DATABASE", "database")
#host = config.get("DATABASE", "host")
#pw = config.get("DATABASE", "password")

try:
    conn = psycopg2.connect(
        "dbname='" + db + "' user='" + user + "' host='"+ host +"' password='" + pw + "'")
except:
    print "I am unable to connect to the database"

cursor = conn.cursor()
cursor_cont = conn.cursor()


query = "SELECT id,ejecutado,estado,consulta FROM crones where nombre ilike 'croncargaA'"
cursor.execute(query)


if cursor.rowcount == 0:
    query = "INSERT INTO crones(nombre, unidad, medida, ejecutado,estado,consulta) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id;"
    cursor.execute(query, ('croncargaA', 1, 'minuto', now.strftime("%Y-%m-%d %H:%M"), 0,0))
else:
    cron = cursor.fetchone()
    if cron[2] == 1:
        
        if cron[3] > 5:
            consulta = 0
            estado = 0
        else:
            consulta = cron[3] + 1
            estado = 1

        query_cont = "update crones set consulta=(%s), estado=(%s), ejecutado=(%s) where id=(%s)"
        cursor.execute(query_cont, (consulta, estado, now.strftime("%Y-%m-%d %H:%M"),cron[0]))
        conn.commit()
        print "Cron en proceso, no se debe ejecutar de nuevo, intento " + str(cron[3])
        exit()
    else:
        query_cont = "update crones set estado=(%s) where id=(%s)"
        cursor.execute(query_cont, (1,cron[0]))
        conn.commit()

conn.commit()   


def getCarrier(data, numero):
    carrie_id = 0
    for row in data:
        if row[2].count(numero[0:3]) > 0:
            carrie_id = row[3]
    return carrie_id

def getPortado(number, carrier_id):
    query = "SELECT current_carrie_id FROM portados where numero = '" + str(number)+"'"
    row = cursor.execute(query)
    if not row:
        return carrier_id
    else:
        print str(number) + " encontro portado "
        return cursor.fetchone()[0]


def getPreference(data_user, carrier_id):
    return data_user[int(carrier_id) - 1]


user_id = 4

query = "SELECT * FROM carries"
cursor.execute(query)
prefijos = cursor.fetchall()

query = "SELECT * FROM usuarios where id = " + str(user_id)
cursor.execute(query)
data_user = cursor.fetchone()
preference_user = data_user[13].split(",")

quantity_files = len(os.listdir(path_files + "/cargados"))

for arc_excel in os.listdir(path_files + "/cargados"):
    
    name_file = path_files + "/cargados/" + arc_excel

# suponiendo que el archivo esta en el mismo directorio del script
    doc = openpyxl.load_workbook(name_file)

    
    
    pesta = doc.sheetnames
    hoja = doc[pesta[0]]

    reg = int(hoja.max_row) - 1

    query = "INSERT INTO bases(idempresa, idusuario, nombre, fecha,archoriginal,estado,registros) VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id;"
    cursor.execute(
        query, (4, user_id, name_file, now.strftime("%Y-%m-%d %H:%M"), name_file,4,reg))
    base_id = cursor.fetchone()[0]
    conn.commit()
    cont=0
    total=0

    for filas in hoja.rows:
        if re.match('\d{10}', str(filas[0].value)) != None:
            print str(filas[0].value)+" "+filas[1].value + " "+filas[2].value
            carrier_id = getCarrier(prefijos, str(filas[0].value))
            if carrier_id > 0:
                carrier_id = getPortado(filas[0].value, carrier_id)
                canal_id = getPreference(preference_user, carrier_id)
                query = "INSERT INTO registrosfast(idbase, numero, mensaje, nota, orden, estado, fechacargue, idcanal, idcarrie, cargue) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
                #query = "INSERT INTO registros(idbase, numero, mensaje, nota, orden, estado, fechacargue, idcanal, idcarrie, cargue) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
                cursor.execute(query, (base_id, str(filas[0].value), filas[1].value, filas[2].value, 1, 2, now.strftime(
                    "%Y-%m-%d %H:%M"), canal_id, carrier_id, 'python'))
            else:
                query = "INSERT INTO errores(idbase, numero, mensaje, nota, orden, estado, fecha, error) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) "
                cursor.execute(query, (base_id, str(filas[0].value), filas[1].value, filas[2].value, 1, 3, now.strftime(
                    "%Y-%m-%d %H:%M"), "Carrier no existe"))
        else:
            query = "INSERT INTO errores(idbase, numero, mensaje, nota, orden, estado, fecha, error) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) "
            cursor.execute(query, (base_id, str(filas[0].value), filas[1].value, filas[2].value, 1, 3, now.strftime(
                "%Y-%m-%d %H:%M"), "No cumple con los requisitos"))

        total = total + 1

        if cont == 10:
            query_cont = "update bases set conteo=(%s) where id=(%s)"
            print query_cont
            cursor.execute(query_cont, (total,base_id))
            conn.commit()        
            cont=0

        cont = cont + 1
        

    query_cont = "update bases set estado=(%s) where id=(%s)"
    cursor.execute(query_cont, (5,base_id))


    query_cont = "update crones set estado=(%s) where id=(%s)"
    cursor.execute(query_cont, (0,cron[0]))

    conn.commit()
    
    conn.close()
    
    print "ejecucion de la base " + str(base_id)
    print "%.2f sec" % (time.clock() - t0)
    
    new = path_files + "/procesados/"+now.strftime("%Y-%m-%d")

    if not os.path.exists(new):
        os.mkdir(new)

    shutil.move(path_files + "/cargados/" + arc_excel, new + "/" + arc_excel)

print "Archivos procesados " + str(quantity_files)


